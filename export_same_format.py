"""Same-format final export helpers for non-bilingual source files."""

from __future__ import annotations

import csv
from io import BytesIO, StringIO

from openpyxl import load_workbook

from export_docx import create_formatted_docx_from_template, fit_text_to_paragraph_count
from export_pdf import create_pdf
from idml_utils import create_translated_idml


EXCEL_EXTENSIONS = {"xlsx", "xlsm"}


def create_same_format_file(source_file_type: str, source_file_bytes: bytes, target_text: str) -> tuple[bytes, str, str]:
    """Create a final file using the same extension as the uploaded input."""
    file_type = (source_file_type or "").lower()
    if not target_text.strip():
        raise ValueError("There is no final target text to export.")

    if file_type == "txt":
        return target_text.encode("utf-8-sig"), "text/plain", ""
    if file_type == "csv":
        return _create_csv(target_text), "text/csv", ""
    if file_type == "pdf":
        return create_pdf("Translation", target_text), "application/pdf", "PDF content is regenerated; original PDF design is not preserved."
    if file_type == "docx":
        return (
            create_formatted_docx_from_template(source_file_bytes, target_text),
            "application/vnd.openxmlformats-officedocument.wordprocessingml.document",
            "DOCX export uses the original DOCX layout template.",
        )
    if file_type == "idml":
        return (
            create_translated_idml(source_file_bytes, target_text),
            "application/vnd.adobe.indesign-idml-package",
            "IDML export preserves the package and replaces editable story text in Stories/*.xml.",
        )
    if file_type in EXCEL_EXTENSIONS:
        mime_type = (
            "application/vnd.ms-excel.sheet.macroEnabled.12"
            if file_type == "xlsm"
            else "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        return (
            create_translated_workbook(source_file_bytes, target_text, keep_vba=file_type == "xlsm"),
            mime_type,
            "Workbook styles and sheets are preserved where possible; text cells are replaced in reading order.",
        )
    if file_type == "xls":
        return (
            _create_excel_compatible_html(target_text),
            "application/vnd.ms-excel",
            "Legacy .xls formatting is not preserved; this creates an Excel-openable .xls table.",
        )
    raise ValueError("Same-format final export is not available for this input type.")


def create_translated_workbook(workbook_bytes: bytes, target_text: str, keep_vba: bool = False) -> bytes:
    """Replace string cells in an XLSX/XLSM workbook while preserving workbook structure."""
    if not workbook_bytes:
        raise ValueError("No original workbook template is available.")

    workbook = load_workbook(BytesIO(workbook_bytes), keep_vba=keep_vba)
    cells = []
    for sheet in workbook.worksheets:
        for row in sheet.iter_rows():
            for cell in row:
                if isinstance(cell.value, str) and cell.value.strip():
                    cells.append(cell)

    if not cells:
        raise ValueError("No editable text cells were found in the workbook.")

    replacements = fit_text_to_paragraph_count(target_text, len(cells))
    for cell, replacement in zip(cells, replacements):
        cell.value = replacement

    output = BytesIO()
    workbook.save(output)
    output.seek(0)
    return output.getvalue()


def _create_csv(target_text: str) -> bytes:
    output = StringIO()
    writer = csv.writer(output)
    for line in [line.strip() for line in target_text.splitlines() if line.strip()]:
        writer.writerow([line])
    return output.getvalue().encode("utf-8-sig")


def _create_excel_compatible_html(target_text: str) -> bytes:
    rows = []
    for line in [line.strip() for line in target_text.splitlines() if line.strip()]:
        escaped = (
            line.replace("&", "&amp;")
            .replace("<", "&lt;")
            .replace(">", "&gt;")
        )
        rows.append(f"<tr><td>{escaped}</td></tr>")
    html = "<html><body><table>" + "".join(rows) + "</table></body></html>"
    return html.encode("utf-8")
