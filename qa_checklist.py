"""Editable QA checklist export/import and approved correction application."""

from __future__ import annotations

from io import BytesIO
from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter

from openai_client import DEFAULT_MODEL, ask_openai


RULE_HEADERS = [
    "ID",
    "Approved",
    "Severity",
    "Category",
    "Issue",
    "Source excerpt",
    "Target excerpt",
    "PM correction / instruction",
]

SUGGESTION_HEADERS = [
    "ID",
    "Apply?",
    "Scope",
    "Suggestion type",
    "Priority",
    "PM-friendly summary",
    "Original QA suggestion",
    "Find / affected wording",
    "Preferred wording",
    "AI action if approved",
    "PM note",
]

LEGACY_SUGGESTION_HEADERS = [
    "ID",
    "Apply?",
    "Scope",
    "Priority",
    "Suggestion",
    "Find / affected wording",
    "Preferred wording",
    "Instruction to AI",
    "PM note",
]

CHECKLIST_HEADERS = RULE_HEADERS
APPROVED_VALUES = {"yes", "y", "true", "1", "approved", "approve", "x", "da"}
HEADER_ROW = 8


def build_qa_checklist_rows(rule_based_warnings: list[dict], qa_report: str) -> list[dict[str, str]]:
    """Return legacy-compatible rows for tests and older callers."""
    return build_rule_based_rows(rule_based_warnings) + [
        _suggestion_as_legacy_row(row) for row in build_suggestion_rows(qa_report)
    ]


def build_rule_based_rows(rule_based_warnings: list[dict]) -> list[dict[str, str]]:
    """Convert deterministic QA warnings into exact correction rows."""
    rows: list[dict[str, str]] = []
    for index, warning in enumerate(rule_based_warnings or [], start=1):
        rows.append(
            {
                "ID": f"R{index}",
                "Approved": "",
                "Severity": str(warning.get("severity") or ""),
                "Category": str(warning.get("category") or "Rule-based QA"),
                "Issue": str(warning.get("message") or ""),
                "Source excerpt": str(warning.get("source excerpt") or ""),
                "Target excerpt": str(warning.get("target excerpt") or ""),
                "PM correction / instruction": "",
                "Type": "rule",
            }
        )
    return rows


def build_suggestion_rows(qa_report: str) -> list[dict[str, str]]:
    """Convert GPT QA notes into PM-friendly global suggestion rows."""
    rows: list[dict[str, str]] = []
    for index, line in enumerate(_qa_report_lines(qa_report), start=1):
        suggestion_type = _suggestion_type(line)
        pm_summary = _pm_friendly_summary(line, suggestion_type)
        rows.append(
            {
                "ID": f"S{index}",
                "Apply?": "",
                "Scope": "Whole document",
                "Suggestion type": suggestion_type,
                "Priority": "Medium",
                "PM-friendly summary": pm_summary,
                "Original QA suggestion": line,
                "Find / affected wording": "",
                "Preferred wording": "",
                "AI action if approved": _default_ai_action(line, suggestion_type),
                "PM note": "",
                "Type": "suggestion",
            }
        )
    return rows


def create_qa_checklist_xlsx(rule_based_warnings: list[dict], qa_report: str) -> bytes:
    """Create an editable Excel checklist for PM approval/rejection."""
    rule_rows = build_rule_based_rows(rule_based_warnings)
    suggestion_rows = build_suggestion_rows(qa_report)
    if not rule_rows and not suggestion_rows:
        raise ValueError("Run QA first before exporting a QA checklist.")

    workbook = Workbook()
    rule_sheet = workbook.active
    _build_sheet(
        rule_sheet,
        "Rule-based corrections",
        "Rule-based QA Corrections",
        [
            "Use this sheet for concrete segment-level issues such as missing numbers, empty targets, URLs, emails, or placeholders.",
            "Set Approved to Yes only when the app should apply this correction.",
            "Use PM correction / instruction when the exact correction needs clarification.",
        ],
        RULE_HEADERS,
        rule_rows,
        widths=[10, 14, 14, 24, 56, 42, 42, 56],
        approval_column="B",
        approval_prompt="Choose Yes only when this exact correction should be applied.",
    )

    suggestion_sheet = workbook.create_sheet("Global suggestions")
    _build_sheet(
        suggestion_sheet,
        "Global suggestions",
        "PM Global Suggestions",
        [
            "Use this sheet for broader style, terminology, register, or consistency improvements.",
            "Set Apply? to Yes only when the suggestion should be applied globally or by the selected scope.",
            "Use Suggestion type and PM-friendly summary to decide quickly; Original QA suggestion keeps the source detail.",
            "Fill Find / affected wording and Preferred wording when you want a terminology or phrase change.",
            "Edit AI action if approved when you want to control exactly how the app applies the suggestion.",
        ],
        SUGGESTION_HEADERS,
        suggestion_rows,
        widths=[10, 14, 22, 20, 14, 42, 58, 34, 34, 58, 36],
        approval_column="B",
        approval_prompt="Choose Yes only when this global suggestion should be applied.",
        scope_column="C",
        priority_column="E",
        suggestion_type_column="D",
    )

    buffer = BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


def read_qa_checklist_xlsx(file_bytes: bytes) -> list[dict[str, str]]:
    """Read an edited QA checklist from Excel."""
    if not file_bytes:
        raise ValueError("Upload a completed QA checklist first.")

    workbook = load_workbook(BytesIO(file_bytes), data_only=True)
    checklist_rows: list[dict[str, str]] = []
    if "Rule-based corrections" in workbook.sheetnames:
        checklist_rows.extend(_read_sheet(workbook["Rule-based corrections"], RULE_HEADERS, "rule"))
    if "Global suggestions" in workbook.sheetnames:
        checklist_rows.extend(_read_suggestion_sheet(workbook["Global suggestions"]))

    if checklist_rows:
        return checklist_rows

    # Backward compatibility with the previous one-sheet checklist.
    return _read_sheet(workbook.active, RULE_HEADERS, "rule")


def approved_checklist_rows(rows: list[dict[str, Any]]) -> list[dict[str, str]]:
    """Return rows the PM approved for AI correction."""
    approved = []
    for row in rows or []:
        value = str(row.get("Approved") or row.get("Apply?") or "").strip().casefold()
        if value in APPROVED_VALUES:
            normalized = {key: str(row.get(key) or "") for key in set(RULE_HEADERS + SUGGESTION_HEADERS + ["Type"])}
            approved.append(normalized)
    return approved


def apply_approved_qa_corrections(
    source_text: str,
    target_text: str,
    approved_rows: list[dict[str, str]],
    target_language: str,
    domain: str,
    model: str = DEFAULT_MODEL,
) -> str:
    """Ask AI to apply only approved QA checklist items to the final translation."""
    if not target_text.strip():
        raise ValueError("There is no final translation to correct.")
    if not approved_rows:
        raise ValueError("No approved QA checklist rows were found. Mark rows as Yes/X/Approved first.")

    checklist = _approved_rows_as_text(approved_rows)
    system_prompt = (
        f"You are a senior translation editor for {target_language}. "
        "Apply approved QA corrections precisely and conservatively."
    )
    user_prompt = f"""Apply the approved QA checklist items to the target translation.

Target language: {target_language}
Text type/domain: {domain}

Rules:
- Apply only approved checklist items.
- Apply rule-based corrections as exact/local corrections.
- Apply approved global suggestions according to their Scope field.
- For global terminology or phrase changes, use Find / affected wording and Preferred wording when supplied.
- Use PM correction / instruction or AI action if approved when supplied.
- Keep the source meaning unchanged.
- Do not introduce corrections from rejected or blank checklist rows.
- Do not rewrite good text unnecessarily.
- Preserve paragraph order, line breaks, numbers, names, dates, units, URLs, emails, placeholders, file markers, and tags.
- Return only the corrected full target translation.

Source text for reference:
{source_text}

Approved QA checklist:
{checklist}

Current target translation:
{target_text}
"""
    corrected = ask_openai(system_prompt, user_prompt, model=model)
    if not corrected.strip():
        raise RuntimeError("AI returned an empty corrected translation.")
    return corrected


def _build_sheet(
    sheet,
    sheet_name: str,
    title: str,
    instructions: list[str],
    headers: list[str],
    rows: list[dict[str, str]],
    widths: list[int],
    approval_column: str,
    approval_prompt: str,
    scope_column: str | None = None,
    priority_column: str | None = None,
    suggestion_type_column: str | None = None,
) -> None:
    """Build one formatted PM checklist sheet."""
    sheet.title = sheet_name
    header_fill = PatternFill("solid", fgColor="E91545")
    header_font = Font(color="FFFFFF", bold=True)
    title_fill = PatternFill("solid", fgColor="111827")
    instruction_fill = PatternFill("solid", fgColor="FFF1F4")
    editable_fill = PatternFill("solid", fgColor="FFF2CC")
    thin_border = Border(
        left=Side(style="thin", color="E5E7EB"),
        right=Side(style="thin", color="E5E7EB"),
        top=Side(style="thin", color="E5E7EB"),
        bottom=Side(style="thin", color="E5E7EB"),
    )

    last_column = len(headers)
    sheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=last_column)
    title_cell = sheet.cell(row=1, column=1, value=title)
    title_cell.fill = title_fill
    title_cell.font = Font(color="FFFFFF", bold=True, size=16)
    title_cell.alignment = Alignment(vertical="center")
    sheet.row_dimensions[1].height = 28

    intro = ["How PMs should use this sheet:"] + instructions
    for offset, text in enumerate(intro, start=2):
        sheet.merge_cells(start_row=offset, start_column=1, end_row=offset, end_column=last_column)
        cell = sheet.cell(row=offset, column=1, value=text)
        cell.fill = instruction_fill
        cell.alignment = Alignment(wrap_text=True, vertical="top")
        cell.font = Font(bold=offset == 2, color="111827")

    for column_index, header in enumerate(headers, start=1):
        cell = sheet.cell(row=HEADER_ROW, column=column_index, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(wrap_text=True, vertical="top")
        cell.border = thin_border

    for row_index, row in enumerate(rows, start=HEADER_ROW + 1):
        for column_index, header in enumerate(headers, start=1):
            cell = sheet.cell(row=row_index, column=column_index, value=row.get(header, ""))
            cell.alignment = Alignment(wrap_text=True, vertical="top")
            cell.border = thin_border
            if header in {
                "Approved",
                "Apply?",
                "PM correction / instruction",
                "AI action if approved",
                "Find / affected wording",
                "Preferred wording",
            }:
                cell.fill = editable_fill
            if header == "Severity":
                _style_severity_cell(cell)

    for column_index, width in enumerate(widths, start=1):
        sheet.column_dimensions[get_column_letter(column_index)].width = width
    for row_index in range(HEADER_ROW + 1, HEADER_ROW + len(rows) + 1):
        sheet.row_dimensions[row_index].height = 64

    _add_list_validation(
        sheet,
        f"{approval_column}{HEADER_ROW + 1}:{approval_column}{HEADER_ROW + max(len(rows), 1)}",
        "Yes,No,Needs discussion",
        approval_prompt,
    )
    if scope_column:
        _add_list_validation(
            sheet,
            f"{scope_column}{HEADER_ROW + 1}:{scope_column}{HEADER_ROW + max(len(rows), 1)}",
            "Whole document,Repeated phrase,Selected segment,Terminology only,Style/register",
            "Choose the scope for this global suggestion.",
        )
    if priority_column:
        _add_list_validation(
            sheet,
            f"{priority_column}{HEADER_ROW + 1}:{priority_column}{HEADER_ROW + max(len(rows), 1)}",
            "High,Medium,Low",
            "Choose the priority for this suggestion.",
        )
    if suggestion_type_column:
        _add_list_validation(
            sheet,
            f"{suggestion_type_column}{HEADER_ROW + 1}:{suggestion_type_column}{HEADER_ROW + max(len(rows), 1)}",
            "Terminology,Style/register,Consistency,Fluency,Accuracy,Formatting,Other",
            "Classify the type of linguistic suggestion.",
        )

    sheet.freeze_panes = f"A{HEADER_ROW + 1}"
    sheet.auto_filter.ref = f"A{HEADER_ROW}:{get_column_letter(last_column)}{HEADER_ROW + max(len(rows), 1)}"


def _read_sheet(sheet, headers: list[str], row_type: str) -> list[dict[str, str]]:
    """Read one formatted or legacy checklist sheet."""
    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        return []
    header_row_index = _find_header_row(rows, headers)
    sheet_headers = [str(value or "").strip() for value in rows[header_row_index]]
    missing = [header for header in headers if header not in sheet_headers]
    if missing:
        raise ValueError(
            "This does not look like a QA checklist exported by the app. "
            f"Missing column(s): {', '.join(missing)}."
        )
    header_index = {header: sheet_headers.index(header) for header in headers}
    checklist_rows = []
    for raw_row in rows[header_row_index + 1:]:
        row = {
            header: _cell_value(raw_row[header_index[header]] if header_index[header] < len(raw_row) else "")
            for header in headers
        }
        row["Type"] = row_type
        if any(value.strip() for value in row.values()):
            checklist_rows.append(row)
    return checklist_rows


def _read_suggestion_sheet(sheet) -> list[dict[str, str]]:
    """Read the current PM-friendly suggestion sheet, with support for older exports."""
    try:
        return _read_sheet(sheet, SUGGESTION_HEADERS, "suggestion")
    except ValueError:
        legacy_rows = _read_sheet(sheet, LEGACY_SUGGESTION_HEADERS, "suggestion")
        return [_modernize_legacy_suggestion_row(row) for row in legacy_rows]


def _find_header_row(rows: list[tuple], headers: list[str]) -> int:
    """Find the checklist header row even when the sheet has instructions above it."""
    if headers == RULE_HEADERS:
        required = {"ID", "Issue"}
    elif headers == LEGACY_SUGGESTION_HEADERS:
        required = {"ID", "Suggestion", "Scope"}
    else:
        required = {"ID", "PM-friendly summary", "Scope"}
    for index, row in enumerate(rows):
        row_headers = {str(value or "").strip() for value in row}
        if required.issubset(row_headers):
            return index
    raise ValueError("This does not look like a QA checklist exported by the app.")


def _add_list_validation(sheet, cell_range: str, values: str, prompt: str) -> None:
    validation = DataValidation(type="list", formula1=f'"{values}"', allow_blank=True)
    validation.error = f"Choose one of: {values}."
    validation.errorTitle = "Invalid value"
    validation.prompt = prompt
    validation.promptTitle = "PM checklist field"
    sheet.add_data_validation(validation)
    validation.add(cell_range)


def _qa_report_lines(qa_report: str) -> list[str]:
    """Use concise non-empty GPT QA report lines as checklist items."""
    lines = []
    for line in str(qa_report or "").splitlines():
        cleaned = line.strip().lstrip("-*0123456789. )\t").strip()
        if len(cleaned) >= 8:
            lines.append(cleaned)
    return lines[:80]


def _suggestion_as_legacy_row(row: dict[str, str]) -> dict[str, str]:
    return {
        "ID": row.get("ID", ""),
        "Approved": row.get("Apply?", ""),
        "Severity": "suggestion",
        "Category": "GPT QA note",
        "Issue": row.get("PM-friendly summary", ""),
        "Source excerpt": "",
        "Target excerpt": "",
        "PM correction / instruction": row.get("AI action if approved", ""),
        "Type": "suggestion",
    }


def _modernize_legacy_suggestion_row(row: dict[str, str]) -> dict[str, str]:
    suggestion = row.get("Suggestion", "")
    suggestion_type = _suggestion_type(suggestion)
    return {
        "ID": row.get("ID", ""),
        "Apply?": row.get("Apply?", ""),
        "Scope": row.get("Scope", ""),
        "Suggestion type": suggestion_type,
        "Priority": row.get("Priority", ""),
        "PM-friendly summary": _pm_friendly_summary(suggestion, suggestion_type) if suggestion else "",
        "Original QA suggestion": suggestion,
        "Find / affected wording": row.get("Find / affected wording", ""),
        "Preferred wording": row.get("Preferred wording", ""),
        "AI action if approved": row.get("Instruction to AI", "") or _default_ai_action(suggestion, suggestion_type),
        "PM note": row.get("PM note", ""),
        "Type": "suggestion",
    }


def _approved_rows_as_text(rows: list[dict[str, str]]) -> str:
    blocks = []
    for row in rows:
        row_type = row.get("Type") or ("suggestion" if row.get("Suggestion") else "rule")
        if row_type == "suggestion":
            instruction = row.get("AI action if approved") or row.get("PM-friendly summary") or row.get("Original QA suggestion") or ""
            blocks.append(
                "\n".join(
                    [
                        f"Type: global suggestion",
                        f"ID: {row.get('ID', '')}",
                        f"Scope: {row.get('Scope', '')}",
                        f"Suggestion type: {row.get('Suggestion type', '')}",
                        f"Priority: {row.get('Priority', '')}",
                        f"PM-friendly summary: {row.get('PM-friendly summary', '')}",
                        f"Original QA suggestion: {row.get('Original QA suggestion', '')}",
                        f"Find / affected wording: {row.get('Find / affected wording', '')}",
                        f"Preferred wording: {row.get('Preferred wording', '')}",
                        f"Approved instruction: {instruction}",
                        f"PM note: {row.get('PM note', '')}",
                    ]
                )
            )
        else:
            instruction = row.get("PM correction / instruction") or row.get("Issue") or ""
            blocks.append(
                "\n".join(
                    [
                        f"Type: rule-based correction",
                        f"ID: {row.get('ID', '')}",
                        f"Severity: {row.get('Severity', '')}",
                        f"Category: {row.get('Category', '')}",
                        f"Issue: {row.get('Issue', '')}",
                        f"Source excerpt: {row.get('Source excerpt', '')}",
                        f"Target excerpt: {row.get('Target excerpt', '')}",
                        f"Approved instruction: {instruction}",
                    ]
                )
            )
    return "\n\n".join(blocks)


def _cell_value(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip()


def _style_severity_cell(cell) -> None:
    """Make severity easy to scan for PMs."""
    severity = str(cell.value or "").casefold()
    if severity == "critical":
        cell.fill = PatternFill("solid", fgColor="FCA5A5")
        cell.font = Font(bold=True, color="7F1D1D")
    elif severity == "warning":
        cell.fill = PatternFill("solid", fgColor="FDE68A")
        cell.font = Font(bold=True, color="78350F")
    elif severity == "suggestion":
        cell.fill = PatternFill("solid", fgColor="DBEAFE")
        cell.font = Font(bold=True, color="1E3A8A")


def _suggestion_type(text: str) -> str:
    """Classify a GPT QA note into a PM-friendly bucket."""
    lowered = text.casefold()
    if any(word in lowered for word in ["term", "terminology", "glossary", "preferred wording"]):
        return "Terminology"
    if any(word in lowered for word in ["register", "tone", "formal", "informal", "style"]):
        return "Style/register"
    if any(word in lowered for word in ["consistent", "consistency", "repeated", "same source"]):
        return "Consistency"
    if any(word in lowered for word in ["fluent", "awkward", "natural", "readability", "phrasing"]):
        return "Fluency"
    if any(word in lowered for word in ["meaning", "mistranslation", "omission", "addition", "accuracy"]):
        return "Accuracy"
    if any(word in lowered for word in ["punctuation", "format", "spacing", "capitalization"]):
        return "Formatting"
    return "Other"


def _pm_friendly_summary(text: str, suggestion_type: str) -> str:
    """Make raw linguistic QA notes easier for PMs to scan."""
    cleaned = " ".join(str(text or "").split())
    prefixes = {
        "Terminology": "Terminology decision needed",
        "Style/register": "Style/register decision needed",
        "Consistency": "Consistency decision needed",
        "Fluency": "Fluency improvement suggested",
        "Accuracy": "Meaning/accuracy check needed",
        "Formatting": "Formatting/punctuation check needed",
        "Other": "Linguistic suggestion",
    }
    return f"{prefixes.get(suggestion_type, 'Linguistic suggestion')}: {cleaned}"


def _default_ai_action(text: str, suggestion_type: str) -> str:
    """Pre-fill an actionable instruction PMs can edit instead of starting blank."""
    cleaned = " ".join(str(text or "").split())
    if suggestion_type == "Terminology":
        return f"Apply this terminology preference consistently where appropriate: {cleaned}"
    if suggestion_type == "Style/register":
        return f"Adjust the affected wording to match the approved style/register: {cleaned}"
    if suggestion_type == "Consistency":
        return f"Harmonize repeated or similar wording according to this suggestion: {cleaned}"
    if suggestion_type == "Fluency":
        return f"Improve fluency only where this issue appears, without changing meaning: {cleaned}"
    if suggestion_type == "Accuracy":
        return f"Correct the target only where this meaning/accuracy issue is present: {cleaned}"
    if suggestion_type == "Formatting":
        return f"Fix formatting, punctuation, spacing, or capitalization according to this suggestion: {cleaned}"
    return f"Apply this approved linguistic suggestion conservatively: {cleaned}"
