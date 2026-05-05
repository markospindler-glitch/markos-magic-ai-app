from __future__ import annotations

import unittest
from io import BytesIO

from openpyxl import Workbook, load_workbook

from export_same_format import create_same_format_file


class ExportSameFormatTests(unittest.TestCase):
    def test_xlsx_same_format_replaces_text_cells(self):
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Source"
        sheet["A1"] = "First source"
        sheet["A2"] = "Second source"
        buffer = BytesIO()
        workbook.save(buffer)

        data, mime_type, note = create_same_format_file(
            "xlsx",
            buffer.getvalue(),
            "First target. Second target.",
        )

        exported = load_workbook(BytesIO(data))
        self.assertEqual(exported.sheetnames, ["Source"])
        self.assertIn("First target", exported["Source"]["A1"].value)
        self.assertIn("Second target", exported["Source"]["A2"].value)
        self.assertIn("spreadsheetml", mime_type)
        self.assertTrue(note)


if __name__ == "__main__":
    unittest.main()
