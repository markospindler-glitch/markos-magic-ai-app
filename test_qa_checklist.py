from __future__ import annotations

import unittest
from io import BytesIO
from unittest.mock import patch

from openpyxl import load_workbook

from qa_checklist import (
    apply_approved_qa_corrections,
    approved_checklist_rows,
    build_suggestion_rows,
    build_qa_checklist_rows,
    create_qa_checklist_xlsx,
    read_qa_checklist_xlsx,
)


class QaChecklistTests(unittest.TestCase):
    def test_build_qa_checklist_rows_from_rule_and_gpt_qa(self):
        rows = build_qa_checklist_rows(
            [
                {
                    "severity": "critical",
                    "category": "Empty target",
                    "message": "Target is empty.",
                    "source excerpt": "Hello.",
                    "target excerpt": "",
                }
            ],
            "Overall result\n- Terminology should be improved.",
        )

        self.assertEqual("R1", rows[0]["ID"])
        self.assertEqual("Empty target", rows[0]["Category"])
        self.assertTrue(any(row["ID"].startswith("S") for row in rows))

    def test_build_suggestion_rows_supports_global_pm_actions(self):
        rows = build_suggestion_rows("Use the client-approved term throughout.")

        self.assertEqual("S1", rows[0]["ID"])
        self.assertEqual("Whole document", rows[0]["Scope"])
        self.assertEqual("Medium", rows[0]["Priority"])
        self.assertIn("client-approved term", rows[0]["Suggestion"])

    def test_qa_checklist_excel_round_trip_reads_approved_rows(self):
        xlsx_bytes = create_qa_checklist_xlsx(
            [
                {
                    "severity": "warning",
                    "category": "Number mismatch",
                    "message": "Review number.",
                    "source excerpt": "Pay 10.",
                    "target excerpt": "Placajte 11.",
                }
            ],
            "",
        )
        workbook = load_workbook(BytesIO(xlsx_bytes))
        sheet = workbook["Rule-based corrections"]
        sheet["B9"] = "Yes"
        sheet["H9"] = "Change the target number to 10."
        edited = BytesIO()
        workbook.save(edited)

        rows = read_qa_checklist_xlsx(edited.getvalue())
        approved = approved_checklist_rows(rows)

        self.assertEqual(1, len(approved))
        self.assertEqual("Yes", approved[0]["Approved"])
        self.assertEqual("Change the target number to 10.", approved[0]["PM correction / instruction"])

    def test_qa_checklist_excel_round_trip_reads_global_suggestions(self):
        xlsx_bytes = create_qa_checklist_xlsx([], "Use more formal register throughout.")
        workbook = load_workbook(BytesIO(xlsx_bytes))
        sheet = workbook["Global suggestions"]
        sheet["B9"] = "Yes"
        sheet["C9"] = "Style/register"
        sheet["H9"] = "Make the translation more formal throughout."
        edited = BytesIO()
        workbook.save(edited)

        rows = read_qa_checklist_xlsx(edited.getvalue())
        approved = approved_checklist_rows(rows)

        self.assertEqual(1, len(approved))
        self.assertEqual("suggestion", approved[0]["Type"])
        self.assertEqual("Style/register", approved[0]["Scope"])
        self.assertEqual("Make the translation more formal throughout.", approved[0]["Instruction to AI"])

    def test_qa_checklist_excel_is_formatted_for_pm_use(self):
        xlsx_bytes = create_qa_checklist_xlsx(
            [
                {
                    "severity": "critical",
                    "category": "Empty target",
                    "message": "Target is empty.",
                    "source excerpt": "Hello.",
                    "target excerpt": "",
                }
            ],
            "",
        )
        workbook = load_workbook(BytesIO(xlsx_bytes))
        sheet = workbook["Rule-based corrections"]

        self.assertEqual("Rule-based QA Corrections", sheet["A1"].value)
        self.assertIn("How PMs should use this sheet", sheet["A2"].value)
        self.assertEqual("Approved", sheet["B8"].value)
        self.assertEqual("A9", sheet.freeze_panes)
        self.assertEqual("A8:H9", sheet.auto_filter.ref)
        self.assertGreaterEqual(len(sheet.data_validations.dataValidation), 1)
        self.assertEqual("critical", sheet["C9"].value)
        self.assertIsNotNone(sheet["C9"].fill.fgColor.rgb)

    def test_qa_checklist_has_global_suggestions_sheet(self):
        xlsx_bytes = create_qa_checklist_xlsx([], "Use approved terminology consistently.")
        workbook = load_workbook(BytesIO(xlsx_bytes))
        sheet = workbook["Global suggestions"]

        self.assertEqual("PM Global Suggestions", sheet["A1"].value)
        self.assertEqual("Apply?", sheet["B8"].value)
        self.assertEqual("Scope", sheet["C8"].value)
        self.assertEqual("Priority", sheet["D8"].value)
        self.assertEqual("Suggestion", sheet["E8"].value)
        self.assertGreaterEqual(len(sheet.data_validations.dataValidation), 3)

    def test_apply_approved_qa_corrections_sends_only_approved_rows(self):
        approved_rows = [
            {
                "ID": "R1",
                "Approved": "Yes",
                "Severity": "critical",
                "Category": "Terminology",
                "Issue": "Wrong term.",
                "Source excerpt": "Source",
                "Target excerpt": "Old target",
                "PM correction / instruction": "Use the approved client term.",
            }
        ]
        with patch("qa_checklist.ask_openai", return_value="Corrected target.") as mocked:
            result = apply_approved_qa_corrections(
                "Source text.",
                "Current target.",
                approved_rows,
                "Slovenian",
                "Legal",
            )

        self.assertEqual("Corrected target.", result)
        prompt = mocked.call_args.args[1]
        self.assertIn("Apply only approved checklist items", prompt)
        self.assertIn("Use the approved client term.", prompt)
        self.assertIn("Current target.", prompt)

    def test_apply_approved_qa_corrections_includes_global_suggestion_scope(self):
        approved_rows = [
            {
                "ID": "S1",
                "Apply?": "Yes",
                "Scope": "Whole document",
                "Priority": "High",
                "Suggestion": "Use preferred term globally.",
                "Find / affected wording": "old term",
                "Preferred wording": "preferred term",
                "Instruction to AI": "Replace the old term wherever appropriate.",
                "PM note": "Client glossary.",
                "Type": "suggestion",
            }
        ]
        with patch("qa_checklist.ask_openai", return_value="Corrected target.") as mocked:
            apply_approved_qa_corrections(
                "Source text.",
                "Current target.",
                approved_rows,
                "Slovenian",
                "Legal",
            )

        prompt = mocked.call_args.args[1]
        self.assertIn("Type: global suggestion", prompt)
        self.assertIn("Scope: Whole document", prompt)
        self.assertIn("preferred term", prompt)


if __name__ == "__main__":
    unittest.main()
